from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtWidgets import * 
from PyQt5.QtGui import * 
from PyQt5.QtWidgets import QApplication, QMainWindow, QGridLayout, QWidget, QTableWidget, QTableWidgetItem
from PyQt5.QtCore import QSize, Qt, QTimer
import sys
import openpyxl
from lib.baselib import Database

class Find_in_Base(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Поиск по базе")
        self.setGeometry(300,250,714,300)
        self.setFixedSize(714,300)

        self.info = QHBoxLayout()

        self.combo_box = QComboBox(self) 
        self.combo_box.move(0, 0)
        self.combo_box.resize(357,25)
        self.chose_list = ["","ФИО", "Место жительства", "Должность", "Оклад"] 
        self.combo_box.addItems(self.chose_list) 
        self.combo_box.setEditable(True)
        self.combo_box.textActivated[str].connect(self.change_name)
        self.info.addWidget(self.combo_box)
        

        
      



        
        self.find_index = QLineEdit(self)
        self.find_index.move(357, 0)
        self.find_index.resize(357,25)
        self.find_index.setPlaceholderText("Выберите вариант")
        self.info.addWidget(self.find_index)

        

        
        self.btn_print_excel = QtWidgets.QPushButton(self)
        self.btn_print_excel.setText("Вывод в Excel")
        self.btn_print_excel.setFixedSize(100,25)
        self.btn_print_excel.move(207,25)
        self.btn_print_excel.clicked.connect(self.print_excel)
        
        
        self.btn_find = QPushButton(self)        
        self.btn_find.setText("Поиск в базе")
        self.btn_find.move(307,25)
        self.btn_find.setFixedSize(100,25)
        self.btn_find.clicked.connect(self.catch_data)




        self.btn_print_base = QtWidgets.QPushButton(self)
        self.btn_print_base.setText("Вывод базы")
        self.btn_print_base.move(407,25)
        self.btn_print_base.setFixedSize(100,25)
        self.btn_print_base.clicked.connect(self.print_base)

        self.table = QTableWidget(self)

        

        self.table.move(0,50)
        self.table.setFixedSize(714,250)
        self.table.setColumnCount(10)
        self.table.resizeColumnToContents(0) 

        self.table.setHorizontalHeaderLabels(["дата","ФИО", "Серия паспорта", "Номер паспорта",
                                               "Дата выдачи", "Кем выдан",
                                                 "Прописка", "Должность",
                                                   "Оклад", "Наличие мед.книжки"])
        
        self.table.horizontalHeaderItem(0).setToolTip("Date")
        self.table.horizontalHeaderItem(1).setToolTip("FIO")
        self.table.horizontalHeaderItem(2).setToolTip("ser_pas")
        self.table.horizontalHeaderItem(3).setToolTip("num_pas")
        self.table.horizontalHeaderItem(4).setToolTip("date_pas")
        self.table.horizontalHeaderItem(5).setToolTip("place_pas")
        self.table.horizontalHeaderItem(6).setToolTip("place_live")
        self.table.horizontalHeaderItem(7).setToolTip("work")
        self.table.horizontalHeaderItem(8).setToolTip("price")
        self.table.horizontalHeaderItem(9).setToolTip("med")
        header = self.table.horizontalHeader()       
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        
        
        self.table.resizeColumnsToContents()






    def print_base(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(10)
        self.check = Database.print_db(self)             
        cnt = self.table.rowCount()
        for i in range(len(self.check)):
            self.table.insertRow(self.table.rowCount())                                        
            for j in range(10):
                item = QTableWidgetItem(self.check[i][j])
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable) 
                if j == 8 and self.check[i][j] == '1':
                    self.table.setItem(cnt+i, j, QTableWidgetItem("+"))
                elif j == 8 and self.check[i][j] == '0':
                    self.table.setItem(cnt+i, j, QTableWidgetItem("-"))
                else:
                    self.table.setItem(cnt+i, j, item)


    def print_excel(self):
        wb = openpyxl.load_workbook("Check.xlsx")
        list = len(wb.get_sheet_names())
        alf = ['A','B','C','D','E','F','G','H','I']
        wb.create_sheet(str(list+1))
        ws = wb[str(list+1)]
        rows = self.table.rowCount()
        for i in range(rows):
            for j in range(len(alf)):
                value = self.table.item(i, j).text()
                ws[alf[j]+str(i+2)].value = value
        wb.save("Check.xlsx")
        QMessageBox.information(self, 'Success', 'Данные успешно сохранены')

    def change_name(self):
        self.txt = self.combo_box.currentText()
        self.find_index.setPlaceholderText('Введите '+self.txt)
    
   

    def catch_data(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(10)
        if self.combo_box.currentText() == 'ФИО':
            self.check = Database.find_FIO(self,self.find_index.text())
            
            if len(self.check) == 0:
                QMessageBox.warning(self, 'Warning', 'Такого человека не найдено')
            else:
                cnt = self.table.rowCount()
                for i in range(len(self.check)):
                    self.table.insertRow(self.table.rowCount())                                        
                    for j in range(10):
                        item = QTableWidgetItem(self.check[i][j])
                        item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable) 
                        if j == 8 and self.check[i][j] == '1':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("+"))
                        elif j == 8 and self.check[i][j] == '0':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("-"))
                        else:
                            self.table.setItem(cnt+i, j, item)
                        

        elif self.combo_box.currentText() == 'Место жительства':
            self.check = Database.find_place_live(self,self.find_index.text())
            
            if len(self.check) == 0:
                QMessageBox.warning(self, 'Warning', 'Такого человека не найдено')
            else:
                cnt = self.table.rowCount()
                for i in range(len(self.check)):
                    self.table.insertRow(self.table.rowCount())                    
                    for j in range(10):
                        item = QTableWidgetItem(self.check[i][j])
                        item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable) 
                        if j == 8 and self.check[i][j] == '1':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("+"))
                        elif j == 8 and self.check[i][j] == '0':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("-"))
                        else:
                            self.table.setItem(cnt+i, j, item)


        elif self.combo_box.currentText() == 'Должность':
            self.check = Database.find_work(self,self.find_index.text())
            if len(self.check) == 0:
                QMessageBox.warning(self, 'Warning', 'Такого человека не найдено')
            else:
                cnt = self.table.rowCount()
                for i in range(len(self.check)):
                    self.table.insertRow(self.table.rowCount())                    
                    for j in range(10):
                        item = QTableWidgetItem(self.check[i][j])
                        item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable) 
                        if j == 8 and self.check[i][j] == '1':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("+"))
                        elif j == 8 and self.check[i][j] == '0':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("-"))
                        else:
                            self.table.setItem(cnt+i, j, item)
        
        elif self.combo_box.currentText() == 'Оклад':
            self.check = Database.find_price(self,self.find_index.text())
            
            if len(self.check) == 0:
                QMessageBox.warning(self, 'Warning', 'Такого человека не найдено')
            else:
                cnt = self.table.rowCount()
                for i in range(len(self.check)):
                    self.table.insertRow(self.table.rowCount())                    
                    for j in range(10):
                        item = QTableWidgetItem(self.check[i][j])
                        item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable) 
                        if j == 8 and self.check[i][j] == '1':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("+"))
                        elif j == 8 and self.check[i][j] == '0':
                            self.table.setItem(cnt+i, j, QTableWidgetItem("-"))
                        else:
                            self.table.setItem(cnt+i, j, item)



        
        

        